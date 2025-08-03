# To check changes
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import uuid
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage

# === STEP 1: LOAD DATA ===
input_csv_path = "D:\\Back up Suraj P\\Downloads\\JIRA (3).csv"
df = pd.read_csv(input_csv_path)

# === STEP 2: DROP UNUSED COLUMNS ===
df.drop(columns=[
    'Assignee Id', 'Reporter Id', 'Priority', 'Due date',
    'Custom field (Comment)', 'Custom field (Title)', 'Reporter', 'Issue id','Custom field (Component).1'
], errors='ignore', inplace=True)

# === STEP 3: REMOVE UNNECESSARY TICKETS ===
df = df[df["Issue Type"] != "[System] Service request"]

# === STEP 4: UPDATE ISSUE TYPE BASED ON MAPPING ===
issue_type_mapping = {
    "[System] Change": "Change Request",
    "[System] Problem": "Problem /Bug",
    "Configuration & data update request": "Configuration & data update request",
    "Data /Report requirement": "Data /Report requirement",
    "Enhancement Request": "Enhancement Request"
}
df["Issue Type"] = df["Issue Type"].replace(issue_type_mapping)

# === STEP 5: CLEANUP AND FORMAT DATES ===
df["Created"] = pd.to_datetime(df["Created"], errors='coerce')
df["Updated"] = pd.to_datetime(df["Updated"], errors='coerce')
df.rename(columns={"Created": "Date"}, inplace=True)
df["Open Days"] = (pd.Timestamp.today() - df["Date"]).dt.days

# === STEP 6: CREATE BLANK FIELDS ===
df["Resolution analysis"] = ""
df["Description (Resolution description)"] = ""

# === STEP 7: MAPPING COLUMNS AS PER SPECIFICATION ===
column_mapping = {
    'Issue Type': 'Issue Type',
    'Issue key': 'Jira no.',
    'Summary': 'Title',
    'Assignee': 'Assignee',
    'Status': 'Status',
    'Resolution': 'Resolution',
    'Date': 'Date',
    'Custom field (Request Source)': 'Request Source',
    'Description': 'Description',
    'Custom field (Component)': 'Component',
    'Custom field (Root causes)': 'Root cause',
    'Custom field (Immediate fix remark)': 'Immediate fix remark',
    'Custom field (Immediate fix target date)': 'Immediate fix target date',
    'Custom field (Immediate fix closure date)': 'Immediate fix closure date',
    'Custom field (Permanent fix required)': 'Permanent fix required',
    'Custom field (Permanent fix remark)': 'Permanent fix remark',
    'Custom field (Issue Occurrence)': 'Issue Occurrence',
    'Custom field (Issue Nature)': 'Issue Nature',
    'Custom field (Issue origin)': 'Issue origin',
    'Custom field (Issue Classification)': 'Type of problem / Issue Classification',
    'Custom field (Severity.)': 'Severity',
    'Custom field (Type of Resolution)': 'Type of Resolution',
    'Custom field (Permanent fix  closure date)': 'Permanent fix closure date',
    'Custom field (Permanent fix  target date)': 'Permanent fix target date',
    'Custom field (Root cause)': 'Root cause'
}

mapped_columns = {col: column_mapping[col] for col in df.columns if col in column_mapping}
df.rename(columns=mapped_columns, inplace=True)
df.columns = df.columns.str.replace(r'^Custom field \((.*?)\)$', r'\1', regex=True)

# === STEP 8: FILTER LAST 30 DAYS ===
last_30_days = pd.Timestamp.today() - pd.Timedelta(days=30)
df_last_30 = df[df["Date"] >= last_30_days]

# === STEP 9: CREATE OUTPUT FILE ===
unique_id = uuid.uuid4().hex[:6]
output_path = f"D:\\Back up Suraj P\\Desktop\\JIRA\\Monthly jira analysis\\Jira_Analysis_with_Summary_{unique_id}.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Cumulative Jira analysis", index=False)
    df_last_30.to_excel(writer, sheet_name="LastMonth_Jira_Analysis", index=False)

    wb = writer.book
    ws_summary = wb.create_sheet("Summary")

    def generate_summary_sheet(dataframe, label_prefix, start_row):
        total_reported = len(dataframe)
        closed_like = dataframe[dataframe["Status"].str.lower().str.contains("resolved|closed|canceled|completed")].shape[0]
        open_mask = ~dataframe["Status"].str.lower().str.contains("resolved|closed|canceled|completed")
        open_count = dataframe[open_mask].shape[0]

        headers = [f"{label_prefix} Total Raised", f"{label_prefix} Total Closed", f"{label_prefix} Total Open"]
        values = [total_reported, closed_like, open_count]

        for col_index, (header, value) in enumerate(zip(headers, values), start=1):
            ws_summary.cell(row=start_row, column=col_index, value=header).font = Font(bold=True)
            ws_summary.cell(row=start_row + 1, column=col_index, value=value)

    generate_summary_sheet(df, "Cumulative", start_row=1)
    generate_summary_sheet(df_last_30, "Last Month", start_row=4)

    # === STEP 10: GRAPHICAL ANALYSIS FOR LAST MONTH ===
    last_month_grouped = df_last_30.copy()
    last_month_grouped["Status Group"] = last_month_grouped["Status"].str.lower().map(
        lambda x: "Open" if not any(term in x for term in ["resolved", "closed", "canceled", "completed"]) else "Closed"
    )
    pivot_chart = last_month_grouped.groupby(["Issue Type", "Status Group"]).size().unstack(fill_value=0).reset_index()

    plt.figure(figsize=(12, 6))
    bar_plot = sns.barplot(
        data=pivot_chart.melt(id_vars='Issue Type', var_name='Status', value_name='Count'),
        x='Issue Type', y='Count', hue='Status', palette='Paired')
    plt.title("Last Month Jira Status by Issue Type")
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    for p in bar_plot.patches:
        height = p.get_height()
        if height > 0:
            bar_plot.annotate(f'{int(height)}', (p.get_x() + p.get_width() / 2, height),
                              ha='center', va='bottom', fontsize=9)

    img_stream = BytesIO()
    plt.savefig(img_stream, format='png')
    plt.close()
    img_stream.seek(0)
    img = XLImage(img_stream)
    img.anchor = "A10"
    ws_summary.add_image(img)

    # === STEP 11: OPEN JIRA AGE BUCKETS ===
    open_issues = df[~df["Status"].str.lower().str.contains("resolved|closed|canceled|completed")]
    bins = [0, 11, 16, 30, 51, np.inf]
    # bins = [0, 10, 15, 30, 50, np.inf]
    labels = ['0-10 days', '11-15 days', '16-30 days', '31-50 days', '>50 days']
    # open_issues['Jira Age Category'] = pd.cut(open_issues['Open Days'], bins=bins, labels=labels, right=True)
    open_issues['Jira Age Category'] = pd.cut(open_issues['Open Days'], bins=bins, labels=labels, right=False)
    age_group_counts = open_issues['Jira Age Category'].value_counts().sort_index().reset_index()
    age_group_counts.columns = ['Jira Age Category', 'Count']

    # Add Aging Table to Summary Sheet
    start_row = 35
    ws_summary.cell(row=start_row - 1, column=1, value="Jira Aging Summary").font = Font(bold=True)
    for r_idx, row in enumerate(dataframe_to_rows(age_group_counts, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws_summary.cell(row=r_idx, column=c_idx, value=value)

    # Add Jira Age Graph
    plt.figure(figsize=(8, 5))
    age_plot = sns.barplot(data=age_group_counts, x='Jira Age Category', y='Count', palette='coolwarm')
    plt.title("Open Jira Ticket Aging Distribution")
    plt.xticks(rotation=45)
    plt.tight_layout()

    for p in age_plot.patches:
        height = p.get_height()
        if height > 0:
            age_plot.annotate(f'{int(height)}', (p.get_x() + p.get_width() / 2, height),
                              ha='center', va='bottom', fontsize=9)

    img_stream2 = BytesIO()
    plt.savefig(img_stream2, format='png')
    plt.close()
    img_stream2.seek(0)
    img2 = XLImage(img_stream2)
    img2.anchor = "A45"
    ws_summary.add_image(img2)

print(f"\u2705 Script execution completed. File saved as: {output_path}")
